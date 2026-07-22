"""
Signature Comparer

Compares two paths (a file or a folder) and dumps the result to Excel:
which files are equal, which changed, which are new/deleted, and which
just got moved around (same hash, different path).
"""

import os
import hashlib
import binascii
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

# Status text used in the report. This is just for internal use (grouping
# rows while building the sheet) - the actual "Status" column in the Excel
# is a formula, not this text directly.
STATUS_EQUAL = "EQUAL"
STATUS_NOT_EQUAL = "NOT EQUAL"
STATUS_NEW = "NEW"
STATUS_DELETED = "DELETED"
STATUS_MOVED = "MOVED"

# Colors for conditional formatting: green/red/yellow, the usual Excel ones.
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

# Header style: bold + a border, no background color.
HEADER_FONT = Font(bold=True)
_MEDIUM = Side(style="medium")
HEADER_BORDER = Border(top=_MEDIUM, bottom=_MEDIUM)
HEADER_BORDER_FIRST_COL = Border(top=_MEDIUM, bottom=_MEDIUM, left=_MEDIUM)


def apply_header_style(ws, row, num_cols):
    """Bold + border for a header row, columns 1 to num_cols."""
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.border = HEADER_BORDER_FIRST_COL if c == 1 else HEADER_BORDER


# Scanning paths and hashing files

def scan_path(root):
    """
    Return {relative_path: full_path} for whatever is in root.

    Works for a single file (one entry, keyed by filename) or a whole folder
    (walked recursively, keyed by path relative to root).
    """
    files = {}
    if os.path.isfile(root):
        files[os.path.basename(root)] = root
    else:
        for dirpath, _dirnames, filenames in os.walk(root):
            for filename in filenames:
                full_path = os.path.join(dirpath, filename)
                rel_path = os.path.relpath(full_path, root)
                files[rel_path] = full_path
    return files


def compute_signatures(filepath, use_sha1, use_crc16, use_md5):
    """
    Hash one file with whatever's enabled.

    Reads the file once in chunks and feeds every enabled hash at the same
    time, so we don't read big files from disk more than once.
    """
    sha1 = hashlib.sha1() if use_sha1 else None
    md5 = hashlib.md5() if use_md5 else None
    crc16 = 0

    with open(filepath, "rb") as f:
        while True:
            chunk = f.read(65536)
            if not chunk:
                break
            if sha1 is not None:
                sha1.update(chunk)
            if md5 is not None:
                md5.update(chunk)
            if use_crc16:
                crc16 = binascii.crc_hqx(chunk, crc16)

    signatures = {}
    if use_sha1:
        signatures["sha1"] = sha1.hexdigest()
    if use_crc16:
        signatures["crc16"] = format(crc16, "04x")
    if use_md5:
        signatures["md5"] = md5.hexdigest()
    return signatures


def signature_key(signatures, hash_order):
    """Turn the signatures dict into a tuple so we can compare/group files easily."""
    return tuple(signatures[h] for h in hash_order)


# Comparison logic

def compare_paths(previous_root, current_root, use_sha1, use_crc16, use_md5):
    """
    Compare previous_root vs current_root, return a list of rows:
    (status, previous_relative_path, current_relative_path,
     previous_signatures, current_signatures)
    plus the list of hash types we actually used (in display order).
    """
    hash_order = []
    if use_sha1:
        hash_order.append("sha1")
    if use_crc16:
        hash_order.append("crc16")
    if use_md5:
        hash_order.append("md5")

    if not hash_order:
        raise ValueError("Select at least one digital signature type.")

    # scan both sides and hash everything
    prev_files = scan_path(previous_root)
    curr_files = scan_path(current_root)

    prev_sigs = {rel: compute_signatures(path, use_sha1, use_crc16, use_md5)
                 for rel, path in prev_files.items()}
    curr_sigs = {rel: compute_signatures(path, use_sha1, use_crc16, use_md5)
                 for rel, path in curr_files.items()}

    common = set(prev_files) & set(curr_files)
    prev_only = set(prev_files) - set(curr_files)
    curr_only = set(curr_files) - set(prev_files)

    rows = []

    # same path on both sides -> equal or not equal, depending on the hash
    for rel in sorted(common):
        prev_key = signature_key(prev_sigs[rel], hash_order)
        curr_key = signature_key(curr_sigs[rel], hash_order)
        status = STATUS_EQUAL if prev_key == curr_key else STATUS_NOT_EQUAL
        rows.append((status, rel, rel, prev_sigs[rel], curr_sigs[rel]))

    # now check what's left (only in previous, or only in current) to catch
    # moved files: same hash, different path = moved, not delete+add
    prev_by_key = {}
    for rel in prev_only:
        key = signature_key(prev_sigs[rel], hash_order)
        prev_by_key.setdefault(key, []).append(rel)

    curr_by_key = {}
    for rel in curr_only:
        key = signature_key(curr_sigs[rel], hash_order)
        curr_by_key.setdefault(key, []).append(rel)

    moved_prev = set()
    moved_curr = set()

    for key, prev_rels in prev_by_key.items():
        curr_rels = curr_by_key.get(key)
        if not curr_rels:
            continue
        # if several files share the same hash, just pair them up in order
        prev_rels_sorted = sorted(prev_rels)
        curr_rels_sorted = sorted(curr_rels)
        for prev_rel, curr_rel in zip(prev_rels_sorted, curr_rels_sorted):
            rows.append((STATUS_MOVED, prev_rel, curr_rel, prev_sigs[prev_rel], curr_sigs[curr_rel]))
            moved_prev.add(prev_rel)
            moved_curr.add(curr_rel)

    # whatever's left in previous with no match = deleted
    for rel in sorted(prev_only - moved_prev):
        rows.append((STATUS_DELETED, rel, "", prev_sigs[rel], {}))

    # whatever's left in current with no match = new
    for rel in sorted(curr_only - moved_curr):
        rows.append((STATUS_NEW, "", rel, {}, curr_sigs[rel]))

    return rows, hash_order


# Excel report

def autofit_columns(ws, min_width=8, padding=2):
    """
    Resize columns to fit their content.

    Skips formula cells on purpose - measuring the formula text itself would
    make columns way too wide (the Status formula is long, but what shows on
    screen is just "EQUAL", "MOVED", etc). Plain text/values still get
    measured normally, so hash and path columns size correctly.
    """
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if value is None:
                continue
            if isinstance(value, str) and value.startswith("="):
                continue
            length = len(str(value))
            col_letter = cell.column_letter
            widths[col_letter] = max(widths.get(col_letter, 0), length)

    for col_letter, length in widths.items():
        ws.column_dimensions[col_letter].width = max(min_width, length + padding)


def write_excel(rows, hash_order, output_path):
    """
    Write everything to one sheet:

      Rows 1-7  : summary (Status / Count / Percentage) for EQUAL, NOT EQUAL,
                  NEW, DELETED, MOVED + a Total row.
      Row 8     : blank row, just a separator.
      Row 9     : header for the detail table (Status, then Previous/Current
                  for each hash type we used, then the two path columns).
      Row 10+   : one row per file.

    The Status column is a formula, not a fixed value - it looks at the
    paths and hashes of its own row and figures out EQUAL/NOT EQUAL/NEW/
    DELETED/MOVED from there. That way if you edit a cell by hand, the
    status (and its color, via conditional formatting) updates itself.
    The summary block up top is also formulas (COUNTIF/SUM) pointing at
    the detail rows below, so it stays in sync too.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"

    # figure out the column layout: Status, then a Previous/Current pair
    # per hash type, then the two path columns
    status_col = 1
    hash_columns = []  # (previous_col, current_col) per hash type
    col = 2
    for _h in hash_order:
        hash_columns.append((col, col + 1))
        col += 2
    prev_path_col = col
    curr_path_col = col + 1
    total_cols = curr_path_col

    prev_path_letter = get_column_letter(prev_path_col)
    curr_path_letter = get_column_letter(curr_path_col)
    last_col_letter = get_column_letter(total_cols)

    # summary goes on top, detail table starts at row 9 (row 8 stays blank)
    detail_header_row = 9
    first_data_row = detail_header_row + 1

    # - detail table first, so we know how many rows we end up with 
    headers = ["Status"]
    for h in hash_order:
        headers.append(f"{h.upper()} Previous")
        headers.append(f"{h.upper()} Current")
    headers.append("Previous Path")
    headers.append("Current Path")
    for c, header in enumerate(headers, start=1):
        ws.cell(row=detail_header_row, column=c, value=header)
    apply_header_style(ws, detail_header_row, len(headers))

    for offset, (_status, prev_rel, curr_rel, prev_sigs, curr_sigs) in enumerate(rows):
        r = first_data_row + offset
        c = 2
        for h in hash_order:
            ws.cell(row=r, column=c, value=prev_sigs.get(h, ""))
            ws.cell(row=r, column=c + 1, value=curr_sigs.get(h, ""))
            c += 2
        ws.cell(row=r, column=prev_path_col, value=prev_rel)
        ws.cell(row=r, column=curr_path_col, value=curr_rel)

        prev_path_ref = f"{prev_path_letter}{r}"
        curr_path_ref = f"{curr_path_letter}{r}"

        # true only if every enabled hash matches between previous and current
        pair_refs = [(get_column_letter(p) + str(r), get_column_letter(cc) + str(r))
                     for p, cc in hash_columns]
        if len(pair_refs) == 1:
            signatures_match = f"{pair_refs[0][0]}={pair_refs[0][1]}"
        else:
            signatures_match = "AND(" + ",".join(f"{p}={cc}" for p, cc in pair_refs) + ")"

        # NEW/DELETED come from a blank path; if both paths are there, it's
        # EQUAL/MOVED/NOT EQUAL depending on the hashes and whether the path
        # itself also matches
        status_formula = (
            f'=IF(AND({prev_path_ref}="",{curr_path_ref}<>""),"{STATUS_NEW}",'
            f'IF(AND({curr_path_ref}="",{prev_path_ref}<>""),"{STATUS_DELETED}",'
            f'IF({signatures_match},'
            f'IF({prev_path_ref}={curr_path_ref},"{STATUS_EQUAL}","{STATUS_MOVED}"),'
            f'"{STATUS_NOT_EQUAL}")))'
        )
        ws.cell(row=r, column=status_col, value=status_formula)

    last_row = first_data_row + len(rows) - 1 if rows else first_data_row

    # color each row based on what the Status formula says
    if rows:
        data_range = f"A{first_data_row}:{last_col_letter}{last_row}"
        ws.conditional_formatting.add(
            data_range, FormulaRule(formula=[f'$A{first_data_row}="{STATUS_EQUAL}"'], fill=FILL_GREEN))
        ws.conditional_formatting.add(
            data_range, FormulaRule(formula=[f'$A{first_data_row}="{STATUS_NOT_EQUAL}"'], fill=FILL_RED))
        ws.conditional_formatting.add(
            data_range, FormulaRule(formula=[f'$A{first_data_row}="{STATUS_NEW}"'], fill=FILL_YELLOW))
        ws.conditional_formatting.add(
            data_range, FormulaRule(formula=[f'$A{first_data_row}="{STATUS_DELETED}"'], fill=FILL_YELLOW))
        ws.conditional_formatting.add(
            data_range, FormulaRule(formula=[f'$A{first_data_row}="{STATUS_MOVED}"'], fill=FILL_YELLOW))

    # - summary block on top, rows 1-7 
    ws.cell(row=1, column=1, value="Status")
    ws.cell(row=1, column=2, value="Count")
    ws.cell(row=1, column=3, value="Percentage")
    apply_header_style(ws, 1, 3)

    statuses = [STATUS_EQUAL, STATUS_NOT_EQUAL, STATUS_NEW, STATUS_DELETED, STATUS_MOVED]
    # COUNTIF just points at the detail rows below (still valid even if there
    # are zero rows, it'll just count 0)
    status_range = f"$A${first_data_row}:$A${last_row}"

    first_count_row = 2
    for offset, status in enumerate(statuses):
        r = first_count_row + offset
        ws.cell(row=r, column=1, value=status)
        ws.cell(row=r, column=2, value=f'=COUNTIF({status_range},"{status}")')

    last_count_row = first_count_row + len(statuses) - 1
    for r in range(first_count_row, last_count_row + 1):
        # avoid a divide-by-zero if there's no data at all
        ws.cell(
            row=r, column=3,
            value=f'=IF(SUM($B${first_count_row}:$B${last_count_row})=0,0,'
                  f'B{r}/SUM($B${first_count_row}:$B${last_count_row}))'
        )
        ws.cell(row=r, column=3).number_format = "0.0%"

    total_row = last_count_row + 1
    ws.cell(row=total_row, column=1, value="Total")
    ws.cell(row=total_row, column=2, value=f"=SUM(B{first_count_row}:B{last_count_row})")
    ws.cell(row=total_row, column=3, value=1)
    ws.cell(row=total_row, column=3).number_format = "0.0%"

    autofit_columns(ws)
    wb.save(output_path)


# GUI

class SignatureComparerApp:
    def __init__(self, root):
        self.root = root
        root.title("Signature comparer")
        root.resizable(False, False)

        self.previous_path = tk.StringVar()
        self.current_path = tk.StringVar()
        self.use_sha1 = tk.BooleanVar(value=True)
        self.use_crc16 = tk.BooleanVar(value=False)
        self.use_md5 = tk.BooleanVar(value=False)

        frame = ttk.Frame(root, padding=15)
        frame.grid(row=0, column=0, sticky="nsew")

        ttk.Label(frame, text="Previous path").grid(row=0, column=0, sticky="w", pady=5)
        ttk.Entry(frame, textvariable=self.previous_path, width=50).grid(row=0, column=1, padx=5)
        ttk.Button(frame, text="Browse",
                   command=lambda: self.browse(self.previous_path)).grid(row=0, column=2)

        ttk.Label(frame, text="Current path").grid(row=1, column=0, sticky="w", pady=5)
        ttk.Entry(frame, textvariable=self.current_path, width=50).grid(row=1, column=1, padx=5)
        ttk.Button(frame, text="Browse",
                   command=lambda: self.browse(self.current_path)).grid(row=1, column=2)

        # SHA1 always on, checkbox just there to show it and stays disabled
        checks = ttk.Frame(frame)
        checks.grid(row=2, column=0, columnspan=3, sticky="w", pady=10)
        ttk.Checkbutton(checks, text="SHA1", variable=self.use_sha1, state="disabled").pack(side="left", padx=5)
        ttk.Checkbutton(checks, text="CRC16", variable=self.use_crc16).pack(side="left", padx=5)
        ttk.Checkbutton(checks, text="MD5", variable=self.use_md5).pack(side="left", padx=5)

        ttk.Button(frame, text="Compare", command=self.on_compare).grid(row=3, column=0, sticky="w", pady=10)

    def browse(self, target_var):
        """Just ask folder or file, then open the matching dialog."""
        is_folder = messagebox.askyesno("Browse", "Do you want to select a folder?\n(No = select a single file)")
        if is_folder:
            path = filedialog.askdirectory()
        else:
            path = filedialog.askopenfilename()
        if path:
            target_var.set(path)

    def on_compare(self):
        previous_root = self.previous_path.get().strip()
        current_root = self.current_path.get().strip()

        if not previous_root or not current_root:
            messagebox.showerror("Error", "Please provide both paths before comparing.")
            return
        if not os.path.exists(previous_root):
            messagebox.showerror("Error", f"The previous path does not exist:\n{previous_root}")
            return
        if not os.path.exists(current_root):
            messagebox.showerror("Error", f"The current path does not exist:\n{current_root}")
            return

        output_path = filedialog.asksaveasfilename(
            title="Save result",
            defaultextension=".xlsx",
            initialfile="comparison_result.xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not output_path:
            return

        try:
            rows, hash_order = compare_paths(
                previous_root, current_root,
                self.use_sha1.get(), self.use_crc16.get(), self.use_md5.get(),
            )
            write_excel(rows, hash_order, output_path)
        except Exception as exc:
            messagebox.showerror("Error", f"An error occurred during the comparison:\n{exc}")
            return

        messagebox.showinfo("Comparison completed", f"Result saved to:\n{output_path}")


def main():
    root = tk.Tk()
    SignatureComparerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
